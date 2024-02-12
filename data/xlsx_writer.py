import xlsxwriter as xw
import statistics
from openpyxl import load_workbook
import os
from tqdm import tqdm
import matplotlib.pyplot as plt
import numpy as np


class WriteExcel:
    def __init__(self, save_path, sample_id: list, recognition_models: list):
        self.end_index = len(sample_id)+2
        self.workbook = xw.Workbook(save_path)
        self.writer_title(sample_id, recognition_models)
        self.bold = self.creat_bold()

    def write(self, data_list, i):
        col_index = chr(ord('B')+i)
        rounded_data_list = [round(num, 4) for num in data_list]
        self.worksheet1.write_column(f"{col_index}2", rounded_data_list, self.bold)  # 从B2单元格开始写入表头
        max_value = max(data_list)
        min_value = min(data_list)
        avg_value = sum(data_list) / len(data_list)
        median_value = statistics.median(data_list)
        variance_value = statistics.variance(data_list)
        self.worksheet1.write(f"{col_index}{self.end_index}", round(max_value, 4), self.bold)
        self.worksheet1.write(f"{col_index}{self.end_index+1}", round(min_value, 4), self.bold)
        self.worksheet1.write(f"{col_index}{self.end_index+2}", round(avg_value, 4), self.bold)
        self.worksheet1.write(f"{col_index}{self.end_index+3}", round(median_value, 4), self.bold)
        self.worksheet1.write(f"{col_index}{self.end_index+4}", round(variance_value, 4), self.bold)

    def close(self):
        self.workbook.close()

    def writer_title(self, sample_id, recognition_models):
        self.worksheet1 = self.workbook.add_worksheet("sheet1")
        self.worksheet1.activate()

        bold = self.workbook.add_format({
            'bold': True,  # 字体加粗
            'border': 3,  # 单元格边框宽度
            'align': 'center',  # 水平对齐方式
            'valign': 'vcenter',  # 垂直对齐方式
            # 'fg_color': '#F4B084',  # 单元格背景颜色
            'text_wrap': True,  # 是否自动换行
        })

        self.worksheet1.set_column('A:E', 15)
        # # demo
        # worksheet1.merge_range('A1:A3', 'Model', bold)
        # worksheet1.write('F2', 'Area=medium', bold)
        # worksheet1.write_row('C3', iou_list, bold)
        self.worksheet1.write_column('A2', sample_id)
        self.worksheet1.write_row('B1', recognition_models)
        self.worksheet1.write(f'A{self.end_index}', 'max', bold)
        self.worksheet1.write(f'A{self.end_index+1}', 'min', bold)
        self.worksheet1.write(f'A{self.end_index+2}', 'avg', bold)
        self.worksheet1.write(f'A{self.end_index+3}', 'med', bold)
        self.worksheet1.write(f'A{self.end_index+4}', 'var', bold)


    def creat_bold(self):
        bold = self.workbook.add_format({
            'bold': True,  # 字体加粗
            'align': 'center',  # 水平对齐方式
        })
        return bold


def statistics_indicators(ckpts):
    result = {}
    colors = ['blue', 'green', 'red', 'purple']
    index_models = ["Facenet512", "SFace", "ArcFace", "VGG-Face"]
    index = ["max", "min", "avg", "med", "var"]
    step = []
    for ckpt in tqdm(ckpts):
        xlsx_file = os.path.join(ckpt, "result.xlsx")
        if not os.path.exists(xlsx_file):
            continue
        wb = load_workbook(xlsx_file)
        sheet = wb.active
        for j, index_ in enumerate(index):
            result.setdefault(index_, {})
            for i, index_model in enumerate(index_models):
                result[index_].setdefault(index_model, [])
                result[index_][index_model].append(sheet[f"{chr(ord('B')+i)}{67+j}"].value)
        
        dir_name = os.path.basename(os.path.dirname(ckpt))
        pretrain_step = int(dir_name.split('step')[-1]) if 'step' in dir_name else 0
        fineturn_step = int(os.path.basename(ckpt).split('-')[-1])
        step.append((pretrain_step+fineturn_step)/1000)
    
    print(f"step:{step[-1]}")
    fig, axs = plt.subplots(len(index), 1, figsize=(20, 20), sharex=True)
    for i, (index_, index_value) in enumerate(result.items()):
        for j, (model, v) in enumerate(index_value.items()):
            axs[i].plot(step, v, color=colors[j], label=model, marker='^')

            if index_=='avg':
                min_index = np.argmin(v)
                min_x = step[min_index]
                min_y = v[min_index]
                axs[i].plot(min_x, min_y, 'k*')
                axs[i].text(min_x, min_y, f'Min:({min_x:.2f}, {min_y:.2f})', verticalalignment='bottom')

        axs[i].set_ylabel(index_)
        axs[i].legend()
    axs[-1].set_xlabel('Step')
    plt.tight_layout()
    save_path = './data/other/eval_model.jpg'
    plt.savefig(save_path)
    print(f"result has saved in {save_path}")


if __name__ == "__main__":
    source_dir = "/mnt/nfs/file_server/public/mingjiahui/experiments/faceid/finetune"
    reset_ckpt_input = [
        f"{source_dir}/base-portrait/20140130-sd15-crop--V1-wo_xformer-scratch",
        f"{source_dir}/base-portrait/20140130-sd15-crop--V1-wo_xformer-scratch_from_step6000",
        f"{source_dir}/base-portrait/20140205-sd15-crop--V1-wo_xformer-scratch_from_step160000",
        f"{source_dir}/base-portrait/20140131-sd15-crop--V1-wo_xformer-scratch_from_step26000/",
        f"{source_dir}/base-portrait/20140205-sd15-crop--V1-wo_xformer-scratch_from_step190000",
        f"{source_dir}/base-portrait/20140205-sd15-crop--V1-wo_xformer-scratch_from_step264000",
    ]
    # reset_ckpt_input = [
    #     f"{source_dir}/portrait-ti_token/20240208-sd15-crop--V1-wo_xformer-scratch/",
    # ]
    # reset_ckpt_input = [
    #     f"{source_dir}/instantid/20240211-sd15--V1-wo_xformer-scratch/",
    # ]
    checkpoint_dirs = []
    for input_dir in reset_ckpt_input:
        checkpoint_dirs += [os.path.join(input_dir, name) for name in os.listdir(input_dir)]
    def extract_number(input):
        dir_name = os.path.basename(os.path.dirname(input))
        pretrain_step = int(dir_name.split('step')[-1]) if 'step' in dir_name else 0
        fineturn_step = int(os.path.basename(input).split('-')[-1])
        return pretrain_step+fineturn_step
    checkpoint_dirs = sorted(checkpoint_dirs, key=extract_number)
    statistics_indicators(checkpoint_dirs)

